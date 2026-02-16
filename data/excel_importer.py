"""
既存 pcf_getter (2).xlsm からのデータ移行モジュール

Excelの各ETFシート（12列 × 93行 + ヘッダー）を読み取り、
PCFRecord のリストに変換する。
"""
from __future__ import annotations

import re
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

from config import (
    EXCEL_PATH,
    EXCEL_SPECIAL_SHEETS,
    TOPIX_ETF_CODES,
    NIKKEI225_ETF_CODES,
)
from models import ETFMaster, PCFRecord, FuturesPosition
from data.parser_futures import normalize_futures

logger = logging.getLogger(__name__)


def _is_special_sheet(sheet_name: str) -> bool:
    """ETFデータではない特殊シートかどうかを判定"""
    # 完全一致
    if sheet_name in EXCEL_SPECIAL_SHEETS:
        return True
    # Chart_ で始まるシートもスキップ
    if sheet_name.startswith("Chart_"):
        return True
    return False


def _extract_etf_code(sheet_name: str) -> Optional[str]:
    """
    シート名からETFコードを抽出する。

    シート名パターン:
      - "1306" (数字のみ)
      - "380A", "399A" (数字+英字)
      - "1306_TOPIX" (コード_名称) の場合はアンダースコア前を取得
    """
    # アンダースコア区切りの場合
    if "_" in sheet_name:
        code_part = sheet_name.split("_")[0].strip()
    else:
        code_part = sheet_name.strip()

    # ETFコードとして妥当か (3-5文字の英数字)
    if re.match(r"^[0-9A-Za-z]{3,6}$", code_part):
        return code_part

    return None


def _to_date(value) -> Optional[date]:
    """Excel セルの値を date に変換"""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            pass
    return None


def _to_float(value) -> Optional[float]:
    """Excel セルの値を float に変換"""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _to_int(value) -> Optional[int]:
    """Excel セルの値を int に変換"""
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _classify_etf(code: str, futures_types: set[str]) -> str:
    """
    ETFをカテゴリに分類する。

    まず config の明示的マッピングを確認し、
    なければ先物データから推定する。
    """
    if code in TOPIX_ETF_CODES:
        return "topix"
    if code in NIKKEI225_ETF_CODES:
        return "nikkei225"

    # 先物データから推定
    if futures_types:
        topix_types = {"TOPIX", "MINI_TOPIX", "TOPIX_BANKS", "TOPIX_CORE30"}
        nk225_types = {"NK225", "NK225_MINI", "NK225_MICRO",
                       "NK225_OPTION_CALL", "NK225_OPTION_PUT"}
        if futures_types & topix_types:
            return "topix"
        if futures_types & nk225_types:
            return "nikkei225"

    return "other"


def import_sheet(ws, etf_code: str) -> list[PCFRecord]:
    """
    1つのワークシートからPCFRecordリストを生成する。

    Args:
        ws: openpyxl ワークシート
        etf_code: ETFコード

    Returns:
        PCFRecord のリスト (日付降順)
    """
    records = []
    is_header = True

    for row in ws.iter_rows(values_only=True):
        # ヘッダー行をスキップ
        if is_header:
            is_header = False
            # ヘッダー行かどうかの確認（最初のセルが文字列ならヘッダー）
            if isinstance(row[0], str) and "日付" in str(row[0]):
                continue
            # ヘッダーでない場合はデータとして処理
            is_header = False

        pcf_date = _to_date(row[0])
        if pcf_date is None:
            continue

        # 先物ポジション
        futures_positions = []

        # 先物1 (Col 6, 7, 8)
        if len(row) > 6 and row[6] is not None:
            raw_name = str(row[6]).strip()
            # ヘッダー文字列が混入するケースを除外
            if raw_name and raw_name not in ("先物の種類1", "先物の種類2"):
                qty = _to_int(row[7]) if len(row) > 7 else 0
                mv = _to_float(row[8]) if len(row) > 8 else 0.0
                fp = normalize_futures(raw_name, qty or 0, mv or 0.0)
                futures_positions.append(fp)

        # 先物2 (Col 9, 10, 11)
        if len(row) > 9 and row[9] is not None:
            raw_name = str(row[9]).strip()
            if raw_name and raw_name not in ("先物の種類1", "先物の種類2"):
                qty = _to_int(row[10]) if len(row) > 10 else 0
                mv = _to_float(row[11]) if len(row) > 11 else 0.0
                fp = normalize_futures(raw_name, qty or 0, mv or 0.0)
                futures_positions.append(fp)

        record = PCFRecord(
            etf_code=etf_code,
            pcf_date=pcf_date,
            nav=_to_float(row[1]),
            shares_outstanding=_to_int(row[2]),
            cash_component=_to_float(row[3]),
            equity_count_tse=_to_int(row[4]) if len(row) > 4 else None,
            equity_market_value=_to_float(row[5]) if len(row) > 5 else None,
            futures_positions=futures_positions,
        )
        records.append(record)

    return records


def import_excel(excel_path: Path = EXCEL_PATH) -> tuple[list[PCFRecord], list[ETFMaster]]:
    """
    既存Excelの全ETFシートからデータを読み込む。

    Args:
        excel_path: Excelファイルパス

    Returns:
        (全PCFRecordリスト, ETFMasterリスト)
    """
    logger.info(f"Excel読み込み開始: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    all_records: list[PCFRecord] = []
    masters: list[ETFMaster] = []
    skipped = []

    for i, sheet_name in enumerate(wb.sheetnames):
        # 特殊シートをスキップ
        if _is_special_sheet(sheet_name):
            skipped.append(sheet_name)
            continue

        etf_code = _extract_etf_code(sheet_name)
        if etf_code is None:
            logger.warning(f"ETFコードを抽出できないシート: '{sheet_name}'")
            skipped.append(sheet_name)
            continue

        ws = wb[sheet_name]
        records = import_sheet(ws, etf_code)

        if records:
            all_records.extend(records)

            # 先物種別を収集
            futures_types = set()
            has_futures = False
            for r in records:
                for fp in r.futures_positions:
                    futures_types.add(fp.futures_type)
                    has_futures = True

            category = _classify_etf(etf_code, futures_types)
            masters.append(ETFMaster(
                code=etf_code,
                name=sheet_name,
                provider="excel_import",
                category=category,
                has_futures=has_futures,
            ))

        if (i + 1) % 50 == 0:
            logger.info(f"  {i + 1}/{len(wb.sheetnames)} シート処理済み")

    wb.close()

    logger.info(
        f"Excel読み込み完了: {len(masters)} ETF, "
        f"{len(all_records)} レコード, "
        f"スキップ: {skipped}"
    )
    return all_records, masters


def records_to_dataframe(records: list[PCFRecord]) -> pd.DataFrame:
    """
    PCFRecordリストをDataFrameに変換する。

    先物ポジションは固定列（先物1, 先物2）として展開する。
    """
    rows = []
    for r in records:
        row = {
            "etf_code": r.etf_code,
            "date": r.pcf_date,
            "nav": r.nav,
            "shares_outstanding": r.shares_outstanding,
            "cash_component": r.cash_component,
            "equity_count_tse": r.equity_count_tse,
            "equity_market_value": r.equity_market_value,
            "nav_per_unit": r.nav_per_unit,
        }

        # 先物1
        if len(r.futures_positions) >= 1:
            fp = r.futures_positions[0]
            row["futures1_raw_name"] = fp.raw_name
            row["futures1_type"] = fp.futures_type
            row["futures1_contract_month"] = fp.contract_month
            row["futures1_quantity"] = fp.quantity
            row["futures1_market_value"] = fp.market_value
            row["futures1_multiplier"] = fp.multiplier
        else:
            row["futures1_raw_name"] = None
            row["futures1_type"] = None
            row["futures1_contract_month"] = None
            row["futures1_quantity"] = None
            row["futures1_market_value"] = None
            row["futures1_multiplier"] = None

        # 先物2
        if len(r.futures_positions) >= 2:
            fp = r.futures_positions[1]
            row["futures2_raw_name"] = fp.raw_name
            row["futures2_type"] = fp.futures_type
            row["futures2_contract_month"] = fp.contract_month
            row["futures2_quantity"] = fp.quantity
            row["futures2_market_value"] = fp.market_value
            row["futures2_multiplier"] = fp.multiplier
        else:
            row["futures2_raw_name"] = None
            row["futures2_type"] = None
            row["futures2_contract_month"] = None
            row["futures2_quantity"] = None
            row["futures2_market_value"] = None
            row["futures2_multiplier"] = None

        rows.append(row)

    df = pd.DataFrame(rows)
    if not df.empty:
        df["date"] = pd.to_datetime(df["date"])
        df = df.sort_values(["etf_code", "date"]).reset_index(drop=True)

    return df


def masters_to_dataframe(masters: list[ETFMaster]) -> pd.DataFrame:
    """ETFMasterリストをDataFrameに変換する"""
    rows = []
    for m in masters:
        rows.append({
            "code": m.code,
            "name": m.name,
            "provider": m.provider,
            "category": m.category,
            "has_futures": m.has_futures,
        })
    return pd.DataFrame(rows)
